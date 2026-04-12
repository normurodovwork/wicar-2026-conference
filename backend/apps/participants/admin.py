from django.contrib import admin
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.urls import path
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.conf import settings
from .models import Participant
import os
import zipfile
import io


@admin.register(Participant)
class ParticipantAdmin(admin.ModelAdmin):
    """Админ-панель для управления участниками."""
    
    list_display = (
        'full_name_display',
        'email_display',
        'affiliation_display',
        'phone_display',
        'article_status',
        'plagiarism_status',
        'payment_status',
        'confirm_payment_link',
        'status_select',
        'download_archive_link',
        'created_at',
    )
    
    list_filter = (
        'status',
        'payment_confirmed',
        'has_article',
        'has_plagiarism',
        'direction',
        'participation_format',
    )
    
    search_fields = (
        'full_name',
        'email',
        'phone',
        'affiliation',
    )
    
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('full_name', 'email', 'phone', 'affiliation')
        }),
        ('Информация из заявки', {
            'fields': ('direction', 'participation_format', 'position', 'talk_type', 'status')
        }),
        ('Файлы', {
            'fields': (
                'has_article', 'article_file',
                'has_plagiarism', 'plagiarism_file',
                'payment_file', 'payment_confirmed'
            )
        }),
        ('Даты', {
            'fields': ('created_at', 'updated_at')
        }),
    )
    
    actions = [
        'confirm_payment',
        'unconfirm_payment',
        'approve_participants',
        'reject_participants',
        'reset_to_pending',
        'export_to_csv',
        'download_selected_archives',
    ]
    
    change_list_template = 'admin/participants/change_list.html'
    
    # Показ статуса
    def full_name_display(self, obj):
        return mark_safe('<strong>{}</strong>'.format(obj.full_name))
    full_name_display.short_description = 'ФИО'
    full_name_display.admin_order_field = 'full_name'
    
    def email_display(self, obj):
        return mark_safe('<a href="mailto:{}">{}</a>'.format(obj.email, obj.email))
    email_display.short_description = 'Email'
    email_display.admin_order_field = 'email'
    
    def affiliation_display(self, obj):
        return obj.affiliation or '-'
    affiliation_display.short_description = 'Учреждение'
    
    def phone_display(self, obj):
        return obj.phone or '-'
    phone_display.short_description = 'Телефон'
    
    def article_status(self, obj):
        if obj.has_article and obj.article_file:
            return mark_safe(
                '<a href="{}" target="_blank" style="color:#4CAF50;font-weight:bold;text-decoration:none;">Скачать</a>'.format(obj.article_file.url)
            )
        return mark_safe('<span style="color:#999;">Нет</span>')
    article_status.short_description = 'Статья'
    
    def plagiarism_status(self, obj):
        if obj.has_plagiarism and obj.plagiarism_file:
            return mark_safe(
                '<a href="{}" target="_blank" style="color:#2196F3;font-weight:bold;text-decoration:none;">Скачать</a>'.format(obj.plagiarism_file.url)
            )
        return mark_safe('<span style="color:#999;">Нет</span>')
    plagiarism_status.short_description = 'Антиплагиат'
    
    def payment_status(self, obj):
        if obj.payment_file:
            return mark_safe(
                format_html(
                    '<img src="{}" class="payment-thumb" onclick="showPaymentModal(\'{}\')" title="Нажмите для увеличения">',
                    obj.payment_file.url, obj.payment_file.url
                )
            )
        if obj.payment_confirmed:
            return mark_safe('<span style="color:#4CAF50;font-weight:bold;">✓ Подтверждено</span>')
        return mark_safe('<span style="color:#999;">Нет оплаты</span>')
    payment_status.short_description = 'Оплата'
    
    def status_select(self, obj):
        """Выпадающий список для изменения статуса прямо в таблице."""
        colors_map = {
            'pending': '#FF9800',
            'approved': '#4CAF50',
            'rejected': '#F44336',
        }
        current_color = colors_map.get(obj.status, '#999')
        html = (
            '<select class="status-select" '
            'onchange="changeStatus({}, this.value, this)" '
            'style="color:{} !important; font-weight:bold !important;">'
            '<option value="pending" {}>На рассмотрении</option>'
            '<option value="approved" {}>Одобрено</option>'
            '<option value="rejected" {}>Отклонено</option>'
            '</select>'
        ).format(
            obj.id,
            current_color,
            'selected' if obj.status == 'pending' else '',
            'selected' if obj.status == 'approved' else '',
            'selected' if obj.status == 'rejected' else '',
        )
        return mark_safe(html)
    status_select.short_description = 'Статус'
    
    def download_archive_link(self, obj):
        """Ссылка для скачивания архива файлов участника."""
        url = f'/admin/participants/participant/{obj.id}/download-archive/'
        return mark_safe(
            f'<a href="{url}" title="Скачать архив" '
            f'style="display:inline-block;width:28px;height:28px;line-height:28px;text-align:center;'
            f'background:#79aec8;color:#fff;border-radius:4px;text-decoration:none;font-size:14px;">⬇</a>'
        )
    download_archive_link.short_description = '📦'
    
    def confirm_payment_link(self, obj):
        """Кнопка подтверждения/отмены оплаты."""
        if obj.payment_confirmed:
            return mark_safe(
                '<button onclick="togglePayment({})" class="btn-action" style="font-size:11px;padding:4px 8px;background:#4CAF50;color:#fff;cursor:pointer;">✓ Оплачено</button>'.format(obj.id)
            )
        return mark_safe(
            '<button onclick="togglePayment({})" class="btn-action" style="font-size:11px;padding:4px 8px;background:#FF9800;color:#fff;cursor:pointer;">Подтвердить</button>'.format(obj.id)
        )
    confirm_payment_link.short_description = 'Оплата'
    
    def get_urls(self):
        """Добавляем кастомные URL."""
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:participant_id>/download-archive/',
                self.admin_site.admin_view(self.download_participant_archive),
                name='participant-download-archive',
            ),
            path(
                'download-all-archives/',
                self.admin_site.admin_view(self.download_all_archives),
                name='participant-download-all-archives',
            ),
            path(
                'export-to-excel/',
                self.admin_site.admin_view(self.export_to_excel),
                name='participant-export-to-excel',
            ),
            path(
                '<int:participant_id>/change-status/',
                self.admin_site.admin_view(self.change_status_ajax),
                name='participant-change-status',
            ),
            path(
                '<int:participant_id>/confirm-payment-inline/',
                self.admin_site.admin_view(self.confirm_payment_inline),
                name='participant-confirm-payment-inline',
            ),
            path(
                '<int:participant_id>/toggle-payment/',
                self.admin_site.admin_view(self.toggle_payment),
                name='participant-toggle-payment',
            ),
        ]
        return custom_urls + urls
    
    def download_participant_archive(self, request, participant_id):
        """Скачивание файлов участника в архиве с ФИО."""
        participant = get_object_or_404(Participant, id=participant_id)
        
        # Создаём имя архива с ФИО
        archive_name = f"{participant.full_name.replace(' ', '_')}_files.zip"
        
        # Создаём архив в памяти
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Добавляем статью
            if participant.article_file and os.path.exists(participant.article_file.path):
                ext = os.path.splitext(participant.article_file.name)[1]
                zf.write(participant.article_file.path, f"Статья{ext}")
            
            # Добавляем антиплагиат
            if participant.plagiarism_file and os.path.exists(participant.plagiarism_file.path):
                ext = os.path.splitext(participant.plagiarism_file.name)[1]
                zf.write(participant.plagiarism_file.path, f"Антиплагиат{ext}")
            
            # Добавляем чек
            if participant.payment_file and os.path.exists(participant.payment_file.path):
                ext = os.path.splitext(participant.payment_file.name)[1]
                zf.write(participant.payment_file.path, f"Чек{ext}")
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{archive_name}"'
        return response
    
    def download_all_archives(self, request):
        """Скачивание всех работ общим архивом (каждый в папке с ФИО)."""
        participants = Participant.objects.all()
        
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for participant in participants:
                # Создаём папку с ФИО
                folder_name = participant.full_name.replace(' ', '_')
                
                # Добавляем статью
                if participant.article_file and os.path.exists(participant.article_file.path):
                    ext = os.path.splitext(participant.article_file.name)[1]
                    arcname = f"{folder_name}/Статья{ext}"
                    zf.write(participant.article_file.path, arcname)
                
                # Добавляем антиплагиат
                if participant.plagiarism_file and os.path.exists(participant.plagiarism_file.path):
                    ext = os.path.splitext(participant.plagiarism_file.name)[1]
                    arcname = f"{folder_name}/Антиплагиат{ext}"
                    zf.write(participant.plagiarism_file.path, arcname)
                
                # Добавляем чек
                if participant.payment_file and os.path.exists(participant.payment_file.path):
                    ext = os.path.splitext(participant.payment_file.name)[1]
                    arcname = f"{folder_name}/Чек{ext}"
                    zf.write(participant.payment_file.path, arcname)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="all_participants_works.zip"'
        return response
    
    def change_status_ajax(self, request, participant_id):
        """AJAX изменение статуса участника."""
        if request.method == 'POST':
            import json
            data = json.loads(request.body)
            participant = get_object_or_404(Participant, id=participant_id)
            new_status = data.get('status')
            
            if new_status in ['pending', 'approved', 'rejected']:
                participant.status = new_status
                participant.save()
                
                status_display = {
                    'pending': 'На рассмотрении',
                    'approved': 'Одобрено',
                    'rejected': 'Отклонено',
                }
                
                return JsonResponse({
                    'success': True,
                    'status': new_status,
                    'status_display': status_display.get(new_status, new_status)
                })
        
        return JsonResponse({'success': False, 'error': 'Invalid request'})
    
    def confirm_payment_inline(self, request, participant_id):
        """AJAX подтверждение оплаты прямо из таблицы."""
        if request.method == 'POST':
            try:
                participant = get_object_or_404(Participant, id=participant_id)
                participant.payment_confirmed = True
                participant.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Оплата подтверждена',
                    'confirmed': True
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=500)
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    def toggle_payment(self, request, participant_id):
        """AJAX переключение статуса оплаты."""
        if request.method == 'POST':
            try:
                participant = get_object_or_404(Participant, id=participant_id)
                participant.payment_confirmed = not participant.payment_confirmed
                participant.save()
                return JsonResponse({
                    'success': True,
                    'confirmed': participant.payment_confirmed,
                    'message': 'Оплата подтверждена' if participant.payment_confirmed else 'Оплата отменена'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=500)
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    def export_to_excel(self, request):
        """Экспорт списка участников в Excel с финансовым отчетом."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.utils import timezone
        from apps.payment.models import PaymentInfo
        
        wb = Workbook()
        
        # ===== ЛИСТ 1: Участники =====
        ws = wb.active
        ws.title = "Участники WICAR 2026"
        
        # Заголовки
        headers = [
            '№', 'ФИО', 'Email', 'Телефон', 'Учреждение',
            'Направление', 'Формат участия', 'Иностранный', 'Должность', 'Тип доклада',
            'Статья', 'Антиплагиат', 'Оплата подтверждена', 'Статус', 'Дата регистрации'
        ]
        
        # Стили
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='417690', end_color='417690', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные
        participants = Participant.objects.all().order_by('created_at')
        for i, p in enumerate(participants, 1):
            row = [
                i,
                p.full_name,
                p.email,
                p.phone or '',
                p.affiliation or '',
                p.direction or '',
                p.participation_format or '',
                'Да' if p.is_foreign else 'Нет',
                p.position or '',
                p.talk_type or '',
                'Да' if p.has_article else 'Нет',
                'Да' if p.has_plagiarism else 'Нет',
                'Да' if p.payment_confirmed else 'Нет',
                p.get_status_display(),
                p.created_at.strftime('%d.%m.%Y %H:%M') if p.created_at else ''
            ]
            
            for col, value in enumerate(row, 1):
                ws.cell(row=i+1, column=col, value=value)
        
        # Ширина колонок
        column_widths = [5, 30, 30, 20, 35, 25, 15, 12, 20, 15, 10, 12, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Автофильтр
        ws.auto_filter.ref = ws.dimensions
        
        # ===== ЛИСТ 2: Финансовый отчет =====
        ws_fin = wb.create_sheet("Финансовый отчет")
        
        # Получаем суммы из настроек
        payment_info = PaymentInfo.objects.filter(is_active=True).first()
        amount_uzs = payment_info.amount_uzs if payment_info else 200000
        amount_usd = payment_info.amount_usd if payment_info else 20
        
        # Заголовки финансового отчета
        fin_headers = [
            '№', 'ФИО', 'Email', 'Учреждение',
            'Иностранный', 'Формат участия',
            'Статус заявки', 'Оплата подтверждена',
            'Сумма взноса', 'Валюта', 'Статус оплаты'
        ]
        
        for col, header in enumerate(fin_headers, 1):
            cell = ws_fin.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные и расчёт сумм
        total_uzs = 0
        total_usd = 0
        confirmed_uzs = 0
        confirmed_usd = 0
        pending_count = 0
        exempt_count = 0
        
        for i, p in enumerate(participants, 1):
            # Определяем сумму
            amount = 0
            currency = 'UZS'
            payment_status = 'Не требуется'
            
            if p.status == 'approved':
                if p.is_foreign and p.participation_format == 'online':
                    # Иностранные онлайн - бесплатно
                    amount = 0
                    currency = '-'
                    payment_status = 'Освобождён'
                    exempt_count += 1
                elif p.is_foreign:
                    amount = amount_usd
                    currency = 'USD'
                    if p.payment_confirmed:
                        payment_status = 'Подтверждено'
                        confirmed_usd += amount
                    else:
                        payment_status = 'Ожидает оплаты'
                        pending_count += 1
                else:
                    amount = amount_uzs
                    currency = 'UZS'
                    if p.payment_confirmed:
                        payment_status = 'Подтверждено'
                        confirmed_uzs += amount
                    else:
                        payment_status = 'Ожидает оплаты'
                        pending_count += 1
            else:
                payment_status = 'Заявка не одобрена'
            
            row = [
                i,
                p.full_name,
                p.email,
                p.affiliation or '-',
                'Да' if p.is_foreign else 'Нет',
                p.participation_format or '-',
                p.get_status_display(),
                'Да' if p.payment_confirmed else 'Нет',
                amount if amount > 0 else '-',
                currency,
                payment_status
            ]
            
            for col, value in enumerate(row, 1):
                ws_fin.cell(row=i+1, column=col, value=value)
        
        # Итоговая строка
        summary_row = len(participants) + 3
        ws_fin.cell(row=summary_row, column=1, value='ИТОГО:')
        ws_fin.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        # Собираем итоги
        summaries = [
            ('Всего участников:', len(participants)),
            ('', ''),
            ('ОПЛАТА UZS:', ''),
            (f'  Подтверждено:', f'{confirmed_uzs:,} UZS'),
            (f'  Ожидает оплаты:', f'{(total_uzs - confirmed_uzs):,} UZS'),
            ('', ''),
            ('ОПЛАТА USD (иностранные):', ''),
            (f'  Подтверждено:', f'{confirmed_usd:,} USD'),
            (f'  Ожидает оплаты:', f'{(total_usd - confirmed_usd):,} USD'),
            ('', ''),
            (f'Иностранцы онлайн (освобождены):', f'{exempt_count} чел.'),
            (f'Ожидает оплаты (всего):', f'{pending_count} чел.'),
        ]
        
        for j, (label, value) in enumerate(summaries):
            ws_fin.cell(row=summary_row + j, column=1, value=label)
            ws_fin.cell(row=summary_row + j, column=1).font = Font(bold=True, size=11)
            ws_fin.cell(row=summary_row + j, column=3, value=value)
        
        # Ширина колонок для финансового отчета
        fin_widths = [5, 30, 30, 35, 12, 15, 18, 15, 15, 10, 18]
        for i, width in enumerate(fin_widths, 1):
            ws_fin.column_dimensions[ws_fin.cell(row=1, column=i).column_letter].width = width
        
        # Автофильтр
        ws_fin.auto_filter.ref = f'A1:K{len(participants) + 1}'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="participants_wicar2026_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    # Actions
    @admin.action(description='Подтвердить оплату выбранных участников')
    def confirm_payment(self, request, queryset):
        queryset.update(payment_confirmed=True)
        self.message_user(request, f'Оплата подтверждена для {queryset.count()} участников')
    
    @admin.action(description='Отменить подтверждение оплаты')
    def unconfirm_payment(self, request, queryset):
        queryset.update(payment_confirmed=False)
        self.message_user(request, f'Подтверждение оплаты отменено для {queryset.count()} участников')
    
    @admin.action(description='Одобрить выбранных участников')
    def approve_participants(self, request, queryset):
        queryset.update(status='approved')
        self.message_user(request, f'{queryset.count()} участников одобрено')
    
    @admin.action(description='Отклонить выбранных участников')
    def reject_participants(self, request, queryset):
        queryset.update(status='rejected')
        self.message_user(request, f'{queryset.count()} участников отклонено')
    
    @admin.action(description='Сбросить статус на "На рассмотрении"')
    def reset_to_pending(self, request, queryset):
        queryset.update(status='pending')
        self.message_user(request, f'Статус сброшен для {queryset.count()} участников')
    
    @admin.action(description='Экспорт в CSV')
    def export_to_csv(self, request, queryset):
        """Экспорт списка участников в CSV."""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="participants.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ФИО', 'Email', 'Телефон', 'Учреждение',
            'Направление', 'Формат', 'Иностранный', 'Должность', 'Тип доклада',
            'Статья', 'Антиплагиат', 'Оплата', 'Статус', 'Дата регистрации'
        ])
        
        for participant in queryset:
            writer.writerow([
                participant.full_name,
                participant.email,
                participant.phone or '',
                participant.affiliation or '',
                participant.direction or '',
                participant.participation_format or '',
                'Да' if participant.is_foreign else 'Нет',
                participant.position or '',
                participant.talk_type or '',
                'Да' if participant.has_article else 'Нет',
                'Да' if participant.has_plagiarism else 'Нет',
                'Да' if participant.payment_confirmed else 'Нет',
                participant.get_status_display(),
                participant.created_at.strftime('%d.%m.%Y %H:%M')
            ])
        
        return response
    
    @admin.action(description='Скачать архивы выбранных участников')
    def download_selected_archives(self, request, queryset):
        """Скачивание архивов выбранных участников."""
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for participant in queryset:
                folder_name = participant.full_name.replace(' ', '_')
                
                if participant.article_file and os.path.exists(participant.article_file.path):
                    ext = os.path.splitext(participant.article_file.name)[1]
                    zf.write(participant.article_file.path, f"{folder_name}/Статья{ext}")
                
                if participant.plagiarism_file and os.path.exists(participant.plagiarism_file.path):
                    ext = os.path.splitext(participant.plagiarism_file.name)[1]
                    zf.write(participant.plagiarism_file.path, f"{folder_name}/Антиплагиат{ext}")
                
                if participant.payment_file and os.path.exists(participant.payment_file.path):
                    ext = os.path.splitext(participant.payment_file.name)[1]
                    zf.write(participant.payment_file.path, f"{folder_name}/Чек{ext}")
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="selected_participants.zip"'
        return response
